"""Tests for the crop -> TAG/DESCRIPTION table path (Claude mocked)."""

import os
import tempfile
import unittest

import app.extraction.claude_api as capi


class TestTagDescriptions(unittest.TestCase):
    def setUp(self):
        self._vc = capi._vision_call

    def tearDown(self):
        capi._vision_call = self._vc

    def test_parses_rows(self):
        capi._vision_call = lambda *a, **k: (
            '[{"tag":"PMP_01","description":"Pump 1 start"},'
            '{"tag":"VLV_A","description":"Valve A"}]')
        rows = capi.tag_descriptions(["img1", "img2"], api_key="sk-x")
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0], {"tag": "PMP_01", "description": "Pump 1 start"})

    def test_empty_on_garbage(self):
        capi._vision_call = lambda *a, **k: "not json"
        self.assertEqual(capi.tag_descriptions(["x"], api_key="sk-x"), [])

    def test_write_tag_table_xlsx(self):
        from app.tools.wizards import _write_tag_table
        from openpyxl import load_workbook
        tmp = tempfile.mkdtemp()
        path = os.path.join(tmp, "tags.xlsx")
        _write_tag_table([{"tag": "A", "description": "alpha"},
                          {"tag": "B", "description": "beta"}], path)
        wb = load_workbook(path)
        rows = list(wb.active.iter_rows(values_only=True))
        self.assertEqual(rows[0], ("TAG", "DESCRIPTION"))
        self.assertEqual(rows[1], ("A", "alpha"))


if __name__ == "__main__":
    unittest.main()
