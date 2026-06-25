"""Tests for crop/extract region classification + Excel/Word/Markdown writers."""

import os
import tempfile
import unittest

import app.extraction.claude_api as capi
from app.tools import region_export as rex


class TestExtractRegionContent(unittest.TestCase):
    def setUp(self):
        self._vc = capi._vision_call

    def tearDown(self):
        capi._vision_call = self._vc

    def test_table_parsed(self):
        capi._vision_call = lambda *a, **k: (
            '{"type":"table","title":"BOM","rows":[["Item","Qty"],["Pump","2"]],'
            '"text":""}')
        out = capi.extract_region_content("pix", api_key="sk-x")
        self.assertEqual(out["type"], "table")
        self.assertEqual(out["title"], "BOM")
        self.assertEqual(out["rows"][1], ["Pump", "2"])

    def test_text_parsed(self):
        capi._vision_call = lambda *a, **k: (
            '{"type":"text","title":"Note","rows":[],"text":"Line 1\\nLine 2"}')
        out = capi.extract_region_content("pix", api_key="sk-x")
        self.assertEqual(out["type"], "text")
        self.assertIn("Line 2", out["text"])

    def test_garbage_returns_empty(self):
        capi._vision_call = lambda *a, **k: "sorry, no json here"
        self.assertEqual(capi.extract_region_content("pix", api_key="sk-x"), {})

    def test_empty_reply_returns_empty(self):
        capi._vision_call = lambda *a, **k: ""
        self.assertEqual(capi.extract_region_content("pix", api_key="sk-x"), {})


def _table(page=0, idx=0, title="BOM"):
    return {"page": page, "index": idx, "type": "table", "title": title,
            "rows": [["Item", "Qty"], ["Pump", "2"], ["Valve", "5"]], "text": ""}


def _text(page=0, idx=1, title="Note"):
    return {"page": page, "index": idx, "type": "text", "title": title,
            "rows": [], "text": "Install per spec.\nTorque to 40 Nm."}


class TestWriters(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp()

    def test_write_excel_tables_only(self):
        path = os.path.join(self.tmp, "t.xlsx")
        self.assertEqual(rex.write_excel([_table(), _text()], path), path)
        from openpyxl import load_workbook
        wb = load_workbook(path)
        self.assertEqual(len(wb.sheetnames), 1)        # only the table region
        ws = wb[wb.sheetnames[0]]
        self.assertEqual(list(next(ws.iter_rows(values_only=True))), ["Item", "Qty"])

    def test_write_excel_none_when_no_tables(self):
        self.assertIsNone(rex.write_excel([_text()], os.path.join(self.tmp, "n.xlsx")))

    def test_write_word_prose_only(self):
        path = os.path.join(self.tmp, "t.docx")
        self.assertEqual(rex.write_word([_table(), _text()], path), path)
        from docx import Document as Docx
        text = "\n".join(p.text for p in Docx(path).paragraphs)
        self.assertIn("Torque to 40 Nm.", text)
        self.assertIn("Note", text)                    # heading
        self.assertNotIn("Pump", text)                 # the table didn't leak in

    def test_write_word_none_when_no_text(self):
        self.assertIsNone(rex.write_word([_table()], os.path.join(self.tmp, "n.docx")))

    def test_write_markdown_everything(self):
        path = os.path.join(self.tmp, "t.md")
        self.assertEqual(rex.write_markdown([_table(), _text()], path), path)
        with open(path) as fh:
            md = fh.read()
        self.assertIn("| Item | Qty |", md)
        self.assertIn("| --- | --- |", md)
        self.assertIn("Torque to 40 Nm.", md)

    def test_export_regions_routes_formats(self):
        written = rex.export_regions(
            [_table(), _text()], self.tmp, "draw",
            formats={"excel", "word", "markdown"})
        names = sorted(os.path.basename(p) for p in written)
        self.assertEqual(names, ["draw_capture.md", "draw_tables.xlsx", "draw_text.docx"])

    def test_export_regions_skips_empty_kinds(self):
        # only text regions -> excel is skipped
        written = rex.export_regions([_text()], self.tmp, "d", formats={"excel", "word"})
        names = [os.path.basename(p) for p in written]
        self.assertEqual(names, ["d_text.docx"])


class TestOcrStructured(unittest.TestCase):
    """Best-effort table reconstruction from positioned OCR words (no Tesseract)."""

    @staticmethod
    def _grid_words():
        # two columns (x~10 and x~210) over three lines -> a table
        words = []
        rows = [("Item", "Qty"), ("Pump", "2"), ("Valve", "5")]
        for r, (a, b) in enumerate(rows):
            top = 10 + r * 30
            words.append({"left": 10, "top": top, "width": 60, "text": a, "line": r})
            words.append({"left": 210, "top": top, "width": 30, "text": b, "line": r})
        return words

    def test_detects_table_columns(self):
        from app.extraction.ocr import words_to_structured
        out = words_to_structured(self._grid_words())
        self.assertEqual(out["type"], "table")
        self.assertEqual(out["rows"][0], ["Item", "Qty"])
        self.assertEqual(out["rows"][2], ["Valve", "5"])

    def test_prose_when_single_column(self):
        from app.extraction.ocr import words_to_structured
        words = [
            {"left": 10, "top": 10, "width": 40, "text": "Install", "line": 0},
            {"left": 55, "top": 10, "width": 30, "text": "per", "line": 0},
            {"left": 90, "top": 10, "width": 40, "text": "spec.", "line": 0},
            {"left": 10, "top": 40, "width": 50, "text": "Torque", "line": 1},
            {"left": 65, "top": 40, "width": 20, "text": "40", "line": 1},
        ]
        out = words_to_structured(words)
        self.assertEqual(out["type"], "text")
        self.assertIn("Install per spec.", out["text"])

    def test_empty_words(self):
        from app.extraction.ocr import words_to_structured
        self.assertEqual(words_to_structured([])["type"], "text")


class TestClassifyRegion(unittest.TestCase):
    def test_ai_path(self):
        import app.tools.wizards as wiz
        orig = capi.extract_region_content
        capi.extract_region_content = lambda *a, **k: {
            "type": "table", "title": "T", "rows": [["a", "b"]], "text": ""}
        try:
            out = wiz._classify_region("pix", 3, 1, True, False, "sk-x", "m")
        finally:
            capi.extract_region_content = orig
        self.assertEqual(out["type"], "table")
        self.assertEqual((out["page"], out["index"]), (3, 1))
        self.assertEqual(out["rows"], [["a", "b"]])

    def test_no_ai_no_ocr_fallback(self):
        import app.tools.wizards as wiz
        out = wiz._classify_region("pix", 0, 0, False, False, "", "m")
        self.assertEqual(out["type"], "text")
        self.assertEqual(out["text"], "")


if __name__ == "__main__":
    unittest.main()
