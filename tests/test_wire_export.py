"""Tests for wire-number export (both modes, separators, labels-per-wire)."""

import csv
import os
import tempfile
import unittest

from app.extraction.wire_parser import WireNumber, TYPE_CONFORMING, TYPE_FIXED, TYPE_JUMPER
from app.export.wire_export import (
    export_single_file, export_per_sheet, WireExportOptions,
    SORT_NUMERICAL, sheet_separator,
)


def _w(label, sheet, rung, idx, page=0, wtype=TYPE_CONFORMING):
    return WireNumber(label=label, sheet=sheet, rung=rung, wire_index=idx,
                      wire_type=wtype, page=page)


class TestWireExport(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.wires = [
            _w("300050", 300, 5, 0), _w("300120", 300, 12, 0),
            _w("301140", 301, 14, 0), _w("301141", 301, 14, 1),
            _w("AB99", None, None, None, wtype=TYPE_FIXED),
            _w("300999", 300, 99, 9, wtype=TYPE_JUMPER),
        ]

    def test_single_file_labels_only_with_separators(self):
        opts = WireExportOptions(fmt="csv", labels_per_wire=2, sort=SORT_NUMERICAL,
                                 include_fixed=False, include_jumpers=False)
        p = export_single_file(self.wires, os.path.join(self.tmp, "w.csv"), opts)
        rows = [r[0] for r in csv.reader(open(p)) if r]
        # separators for sheet 300 and 301
        self.assertIn("~300~", rows)
        self.assertIn("~301~", rows)
        # each conforming label twice
        self.assertEqual(rows.count("300050"), 2)
        self.assertEqual(rows.count("301141"), 2)
        # excluded types absent
        self.assertNotIn("AB99", rows)
        self.assertNotIn("300999", rows)

    def test_single_file_default_is_labels_only(self):
        opts = WireExportOptions(fmt="csv", labels_per_wire=1)
        p = export_single_file(self.wires, os.path.join(self.tmp, "w2.csv"), opts)
        rows = list(csv.reader(open(p)))
        # no header row like "Label,Sheet,..."
        self.assertNotIn("Label", [r[0] for r in rows if r])

    def test_per_sheet_full_columns(self):
        opts = WireExportOptions(fmt="csv", labels_per_wire=1, include_jumpers=False)
        paths = export_per_sheet(self.wires, os.path.join(self.tmp, "out"), opts)
        names = sorted(os.path.basename(p) for p in paths)
        self.assertIn("WIRES_300.csv", names)
        self.assertIn("WIRES_301.csv", names)
        # full columns by default in per-sheet mode
        rows = list(csv.reader(open(os.path.join(self.tmp, "out", "WIRES_300.csv"))))
        self.assertEqual(rows[0], ["Label", "Sheet", "Rung", "WireIdx", "Type", "Page", "Source"])

    def test_xlsx_single_file(self):
        from openpyxl import load_workbook
        opts = WireExportOptions(fmt="xlsx", labels_per_wire=1, include_fixed=False,
                                 include_jumpers=False)
        p = export_single_file(self.wires, os.path.join(self.tmp, "w.xlsx"), opts)
        wb = load_workbook(p)
        vals = [c[0] for c in wb.active.iter_rows(values_only=True)]
        self.assertIn("~300~", vals)

    def test_separator_format(self):
        self.assertEqual(sheet_separator(201), "~201~")
        self.assertEqual(sheet_separator(None), "~?~")

    def test_only_included_toggle(self):
        self.wires[0].included = False
        opts = WireExportOptions(fmt="csv", labels_per_wire=1, only_included=True)
        p = export_single_file(self.wires, os.path.join(self.tmp, "w3.csv"), opts)
        rows = [r[0] for r in csv.reader(open(p)) if r]
        self.assertNotIn("300050", rows)


if __name__ == "__main__":
    unittest.main()
